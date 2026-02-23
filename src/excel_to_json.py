# --------------------------------------------------
# import
# --------------------------------------------------
import openpyxl
from pathlib import Path
import json

from logger import LOGGER
from define import FULLPATH_ORIGINAL_EXCEL, FULLPATH_INTERMEDIATE_JSON, EXCEL_EXTENSIONS
from basic_func import create_dir, get_file_path_recurcive_as_list
from data_structure_cell import CellData


# --------------------------------------------------
# class
# --------------------------------------------------
class ExcelToJson:
    def __init__(self) -> None:
        LOGGER.info(f"開始します")
        LOGGER.debug(f"Excelファイル格納ディレクトリ: {FULLPATH_ORIGINAL_EXCEL}")
        LOGGER.debug(f"JSONファイル格納ディレクトリ: {FULLPATH_INTERMEDIATE_JSON}")

        # JSONファイル格納ディレクトリを作成
        create_dir(FULLPATH_INTERMEDIATE_JSON)

        # Excelファイル格納ディレクトリ内の全てのファイルをJSONファイルに変換して保存
        excel_path_list = get_file_path_recurcive_as_list(FULLPATH_ORIGINAL_EXCEL, EXCEL_EXTENSIONS)
        for excel_path in excel_path_list:
            self.convert_excel_to_json(excel_path)

        LOGGER.info(f"終了しました")

    def convert_excel_to_json(self, excel_path: str) -> None:
        """
        指定されたExcelファイルの全シートをJSONに変換して保存する。
        出力ファイル名は「{元ファイル名}.json」。
        """
        LOGGER.info(f"処理します: {excel_path}")

        result = {}

        # Excelファイルを読み込む
        wb = openpyxl.load_workbook(excel_path, data_only=False)

        # 全シートを舐める
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheet_data = []
            # 全行を舐める
            for row in ws.iter_rows():
                # 全セルを舐める
                for cell in row:
                    # 空セルはスキップ
                    if cell.value is None:
                        continue
                    # セルデータ抽出
                    cell_data = CellData(cell)
                    cell_data_dict = cell_data.to_dict()
                    LOGGER.debug(f"セルデータ: {cell_data_dict}")
                    # セルデータ格納
                    sheet_data.append(cell_data_dict)

            result[sheet] = sheet_data

        # JSON保存
        output_file_path = f"{FULLPATH_INTERMEDIATE_JSON}/{Path(excel_path).stem}.json"
        with open(output_file_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=4)

        LOGGER.info(f"保存しました: {output_file_path}")


# --------------------------------------------------
# main
# --------------------------------------------------
if __name__ == "__main__":
    ExcelToJson()
